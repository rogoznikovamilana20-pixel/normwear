"""
Parse МойСклад Excel export → clean JSON for import.
Structure: Товар row = parent product, followed by Модификация rows = color/size variants.
"""
import openpyxl, json, re
from pathlib import Path
from collections import defaultdict

INPUT = Path(r'C:\Users\andre\Downloads\Telegram Desktop\20260902150228_admin@bazhukovsergey_665440570.xlsx')
OUTPUT = Path(__file__).parent.parent / 'moysklad_excel_products.json'

wb = openpyxl.load_workbook(INPUT, data_only=True)
ws = wb['Sheet0']

# Column indices (0-based from headers.txt)
COL_GROUP = 0
COL_UUID = 1
COL_TYPE = 2
COL_CODE = 3
COL_NAME = 4
COL_ARTICLE = 6
COL_PRICE_RETAIL = 8
COL_PRICE_WHOLESALE = 10
COL_PRICE_SITE = 12
COL_PURCHASE = 16
COL_SUPPLIER = 34
COL_PARENT_CODE = 36
COL_ARCHIVE = 37
COL_GENDER = 49
COL_IMAGE = 54
COL_COLOR = 55
COL_SIZE = 56

def parse_price(val):
    """Parse '4300,00' → 4300.0"""
    if val is None:
        return 0.0
    s = str(val).replace(',', '.').replace(' ', '').replace('\xa0', '')
    try:
        return float(s)
    except ValueError:
        return 0.0

# First pass: collect all Товар rows as product headers
products = {}  # code → product dict
current_product = None

for r in range(2, ws.max_row + 1):
    row_type = ws.cell(r, COL_TYPE + 1).value
    group = ws.cell(r, COL_GROUP + 1).value
    code = ws.cell(r, COL_CODE + 1).value
    name = ws.cell(r, COL_NAME + 1).value
    image = ws.cell(r, COL_IMAGE + 1).value
    retail = ws.cell(r, COL_PRICE_RETAIL + 1).value
    wholesale = ws.cell(r, COL_PRICE_WHOLESALE + 1).value
    site_price = ws.cell(r, COL_PRICE_SITE + 1).value
    purchase = ws.cell(r, COL_PURCHASE + 1).value
    color = ws.cell(r, COL_COLOR + 1).value
    size = ws.cell(r, COL_SIZE + 1).value
    uuid = ws.cell(r, COL_UUID + 1).value

    if row_type == 'Товар':
        # Parse group: "Куртки/Oakley" → category="Куртки", brand="Oakley"
        category = ''
        brand = ''
        if group:
            parts = group.split('/')
            category = parts[0].strip()
            if len(parts) > 1:
                brand = parts[1].strip()

        code_str = str(code).strip() if code else ''
        current_product = {
            'code': code_str,
            'uuid': str(uuid) if uuid else '',
            'name': str(name).strip() if name else '',
            'category': category,
            'brand': brand,
            'retail_price': parse_price(retail),
            'wholesale_price': parse_price(wholesale),
            'site_price': parse_price(site_price),
            'purchase_price': parse_price(purchase),
            'image_url': str(image).strip() if image and str(image).startswith('http') else '',
            'sizes': [],
            'colors': [],
            'mods': [],
        }
        products[code_str] = current_product

    elif row_type == 'Модификация' and current_product:
        mod = {
            'code': str(ws.cell(r, COL_CODE + 1).value or '').strip(),
            'uuid': str(uuid) if uuid else '',
            'color': str(color).strip() if color else '',
            'size': str(size).strip() if size else '',
            'retail_price': parse_price(ws.cell(r, COL_PRICE_RETAIL + 1).value),
            'purchase_price': parse_price(ws.cell(r, COL_PURCHASE + 1).value),
        }
        current_product['mods'].append(mod)
        if mod['size'] and mod['size'] not in current_product['sizes']:
            current_product['sizes'].append(mod['size'])
        if mod['color'] and mod['color'] not in current_product['colors']:
            current_product['colors'].append(mod['color'])

wb.close()

# Clean up and compute final fields
result = []
for code, p in products.items():
    if not p['name'] or p['name'] == 'ИЗ':
        continue  # skip service entries

    # Build description from brand + category + colors
    desc_parts = []
    if p['brand']:
        desc_parts.append(p['brand'])
    if p['colors']:
        desc_parts.append(f"Цвета: {', '.join(p['colors'])}")
    description = ' | '.join(desc_parts) if desc_parts else p['brand']

    # Build title
    title = f"{p['name']} [{p['code']}]" if p['code'] else p['name']

    # Determine retail price: if parent is 0, use first mod price
    retail = p['retail_price']
    if retail <= 0 and p['mods']:
        for m in p['mods']:
            if m['retail_price'] > 0:
                retail = m['retail_price']
                break

    # Sort sizes
    size_order = {'XXS': 0, 'XS': 1, 'S': 2, 'M': 3, 'L': 4, 'XL': 5, '2XL': 6, '3XL': 7, '4XL': 8}
    p['sizes'].sort(key=lambda s: size_order.get(s.upper(), 99))

    result.append({
        'code': p['code'],
        'uuid': p['uuid'],
        'title': title,
        'name': p['name'],
        'category': p['category'],
        'brand': p['brand'],
        'description': description,
        'retail_price': retail,
        'wholesale_price': p['wholesale_price'],
        'site_price': p['site_price'],
        'purchase_price': p['purchase_price'],
        'image_url': p['image_url'],
        'sizes': p['sizes'],
        'colors': p['colors'],
        'mod_count': len(p['mods']),
        'mods': p['mods'],
    })

# Sort by category, then name
result.sort(key=lambda x: (x['category'], x['brand'], x['name']))

# Stats
with_image = sum(1 for p in result if p['image_url'])
with_price = sum(1 for p in result if p['retail_price'] > 0)
total_mods = sum(p['mod_count'] for p in result)

print(f"Products: {len(result)}")
print(f"With image: {with_image}")
print(f"With price > 0: {with_price}")
print(f"Total variants: {total_mods}")
print(f"\nCategories:")
cats = defaultdict(int)
for p in result:
    cats[f"{p['category']}/{p['brand']}"] += 1
for c, n in sorted(cats.items(), key=lambda x: -x[1])[:20]:
    print(f"  {c}: {n}")

with open(OUTPUT, 'w', encoding='utf-8') as f:
    json.dump(result, f, ensure_ascii=False, indent=2)
print(f"\nSaved: {OUTPUT}")
